import asyncio
import functools
import inspect
import json
import logging
import os
import time
import traceback
from copy import (
    copy,
)
from tempfile import (
    NamedTemporaryFile,
)

import requests
from openpyxl import (
    Workbook,
    load_workbook,
)
from openpyxl.cell import (
    WriteOnlyCell,
)
from openpyxl.styles.named_styles import (
    NamedStyle,
)

from connect.client import (
    AsyncConnectClient,
    ConnectClient,
)
from connect.client.models import (
    AsyncResource,
)
from connect.eaas.core.enums import (
    ResultType,
)
from connect.eaas.core.proto import (
    Task,
    TaskOutput,
)
from connect.eaas.core.responses import (
    RowTransformationResponse,
    TransformationResponse,
)
from connect.eaas.runner.constants import (
    DOWNLOAD_CHUNK_SIZE,
    EXCEL_NULL_MARKER,
    REPORT_EVERY_ROW_MAX,
    ROW_DELETED_MARKER,
    TRANSFORMATION_TASK_MAX_PARALLEL_LINES,
    UPLOAD_CHUNK_SIZE,
)
from connect.eaas.runner.logging import (
    RequestLogger,
)
from connect.eaas.runner.managers.base import (
    TasksManagerBase,
)


logger = logging.getLogger(__name__)


class RowTransformationError(Exception):
    pass


class TransformationTasksManager(TasksManagerBase):
    """
    Class for Transformations managers.
    """

    def get_client(self, task_data):
        if task_data.options.api_key:
            return self._task_api_key_clients.setdefault(
                task_data.options.api_key,
                AsyncConnectClient(
                    task_data.options.api_key,
                    endpoint=self.config.get_api_url(),
                    use_specs=False,
                    default_headers=self.config.get_user_agent(),
                    logger=RequestLogger(logger),
                ),
            )

        return self.client

    def get_extension_logger(self, task_data):
        return self.handler.get_logger(
            extra={'task_id': task_data.options.task_id},
        )

    def get_sync_client(self, task_data):
        return ConnectClient(
            task_data.options.api_key,
            endpoint=self.config.get_api_url(),
            use_specs=False,
            default_headers=self.config.get_user_agent(),
            logger=RequestLogger(logger),
        )

    def send_skip_response(self, data, output):
        future = asyncio.Future()
        future.set_result(TransformationResponse.skip(output))
        asyncio.create_task(self.enqueue_result(data, future))

    async def get_argument(self, task_data):
        """
        Get the transformation request through Connect public API
        related to the processing task.
        """
        client = self.get_client(task_data)
        object_exists = await self.filter_collection_by_event_definition(
            client,
            task_data,
        )
        if not object_exists:
            return

        definition = self.config.event_definitions[task_data.input.event_type]
        url = definition.api_resource_endpoint.format(pk=task_data.input.object_id)
        resource = AsyncResource(client, url)

        return await resource.get()

    def get_method_name(self, task_data, argument):
        return task_data.input.data['method']

    async def invoke(self, task_data, method, tfn_request):
        """
        Creates async task for transformation process.

        :param task_data: Data of the task to be processed.
        :type task_data: connect.eaas.core.proto.Task
        :param method: The method that has to be invoked.
        :param tfn_request: Connect tfn request that need to be processed.
        :type tfn_request: dict
        """

        future = asyncio.create_task(self.process_transformation(
            task_data, tfn_request, method,
        ))

        logger.info(f'Enqueue result for task {task_data.options.task_id}')
        asyncio.create_task(self.enqueue_result(task_data, future))

    async def build_response(self, task_data, future):
        result_message = Task(**task_data.dict())
        timeout = self.config.get_timeout('transformation')
        try:
            begin_ts = time.monotonic()
            result = await asyncio.wait_for(
                future,
                timeout=timeout,
            )
            result_message.output = TaskOutput(result=result.status)
            result_message.output.runtime = time.monotonic() - begin_ts
            logger.info(
                f'Transformation task {task_data.options.task_id} '
                f'result: {result.status}, '
                f'took: {result_message.output.runtime}',
            )
            if result.status in (ResultType.SKIP, ResultType.FAIL):
                result_message.output.message = result.output
        except Exception as e:
            cause = (
                str(e) if not isinstance(e, asyncio.TimeoutError)
                else f'timed out after {timeout} s'
            )
            self.log_exception(task_data, e)
            await self._fail_task(task_data, cause)
            result_message.output = TaskOutput(result=ResultType.FAIL)
            result_message.output.message = traceback.format_exc()[:4000]

        return result_message

    async def _fail_task(self, task_data, message):
        try:
            client = self.get_client(task_data)
            ns, _ = task_data.input.event_type.split('_', 1)
            await client(ns).requests[task_data.input.object_id]('fail').post()
            await client.conversations[task_data.input.object_id].messages.create(
                payload={
                    'type': 'message',
                    'text': f'Transformation request processing failed: {message}.',
                },
            )
        except Exception:
            logger.exception(f'Cannot fail the transformation request {task_data.input.object_id}')

    async def process_transformation(self, task_data, tfn_request, method):
        extension_logger = self.get_extension_logger(task_data)
        input_file = await asyncio.get_running_loop().run_in_executor(
            self.executor,
            self.download_excel,
            tfn_request,
            task_data.options.api_key,
            extension_logger,
        )
        output_file = NamedTemporaryFile(
            suffix=f'.{tfn_request["files"]["input"]["name"].split(".")[-1]}',
        )

        read_queue = asyncio.Queue(TRANSFORMATION_TASK_MAX_PARALLEL_LINES)
        write_queue = asyncio.Queue()

        loop = asyncio.get_event_loop()

        reader_task = loop.run_in_executor(
            self.executor,
            self.read_excel,
            tfn_request,
            input_file,
            read_queue,
            extension_logger,
            loop,
        )
        writer_task = loop.run_in_executor(
            self.executor,
            self.write_excel,
            output_file.name,
            write_queue,
            tfn_request['stats']['rows']['total'],
            tfn_request['transformation']['columns']['output'],
            task_data,
            extension_logger,
            loop,
        )
        processor_task = asyncio.create_task(self.process_rows(
            read_queue,
            write_queue,
            method,
            tfn_request,
            extension_logger,
        ))

        tasks = [reader_task, writer_task, processor_task]
        try:
            await asyncio.gather(*tasks)
        except Exception as e:
            logger.exception(
                f'Error during processing transformation '
                f'for {task_data.options.task_id}: {e}',
            )
            for task in tasks:
                if not task.done():
                    task.cancel()
            input_file.close()
            output_file.close()
            client = self.get_client(task_data)
            ns, _ = task_data.input.event_type.split('_', 1)
            await client(ns).requests[task_data.input.object_id]('fail').post()
            await client.conversations[task_data.input.object_id].messages.create(
                payload={
                    'type': 'message',
                    'text': (
                        'Transformation request processing failed: '
                        f'{self.format_exception_message(e)}'
                    ),
                },
            )
            return TransformationResponse.fail(output=str(e))

        await self.send_output_file(
            task_data, tfn_request['batch']['id'], output_file, extension_logger,
        )
        input_file.close()
        output_file.close()
        return TransformationResponse.done()

    def download_excel(self, tfn_request, api_key, logger):
        input_file_name = tfn_request['files']['input']['name']
        input_file = NamedTemporaryFile(suffix=f'.{input_file_name.split(".")[-1]}')
        logger.info(
            f'Downloading input file for {tfn_request["id"]} '
            f'from {self.config.get_api_address()}{input_file_name}',
        )
        with requests.get(
            url=f'{self.config.get_api_address()}{input_file_name}',
            stream=True,
            headers={
                'Authorization': api_key,
                **self.config.get_user_agent(),
            },
        ) as response:
            response.raise_for_status()
            content_length = response.headers.get('Content-Length')
            progress = 0
            for chunk in response.iter_content(chunk_size=DOWNLOAD_CHUNK_SIZE):
                input_file.write(chunk)
                progress += len(chunk)
                logger.debug(
                    f'Input file download progress for {tfn_request["id"]}:'
                    f' {progress}/{content_length} bytes',
                )
        logger.info(
            f'Input file for {tfn_request["id"]} '
            f'from {self.config.get_api_address()}{input_file_name} downloaded',
        )
        return input_file

    def read_excel(self, tfn_request, filename, queue, logger, loop):
        wb = load_workbook(filename=filename, read_only=True)
        ws = wb['Data']
        lookup_columns = {}
        total_rows = tfn_request['stats']['rows']['total']
        delta = 1 if total_rows <= 10 else round(total_rows / 10)
        for idx, row in enumerate(ws.rows, start=1):
            if idx == 1:
                for col_idx, col_value in enumerate(row, start=1):
                    lookup_columns[col_idx] = col_value.value
                continue

            row_data = {col_name: None for col_name in lookup_columns.values()}
            row_styles = copy(row_data)

            for col_idx, col_value in enumerate(row, start=1):
                row_data[lookup_columns[col_idx]] = col_value.value
                style = NamedStyle()
                style.font = col_value.font
                style.fill = col_value.fill
                style.border = col_value.border
                style.number_format = col_value.number_format
                row_styles[lookup_columns[col_idx]] = style

            future = asyncio.run_coroutine_threadsafe(
                queue.put((idx, row_data, row_styles)),
                loop,
            )
            future.result()
            if idx % delta == 0 or idx == total_rows:
                logger.info(
                    f'Input file read progress for {tfn_request["id"]}:'
                    f' {idx}/{total_rows} rows',
                )

        logger.info(f'Input file read complete for {tfn_request["id"]}')
        wb.close()

    async def process_rows(self, read_queue, result_store, method, tfn_request, logger):
        rows_processed = 0
        tasks = []
        total_rows = tfn_request['stats']['rows']['total']
        delta = 1 if total_rows <= 10 else round(total_rows / 10)
        while rows_processed < total_rows:
            row_idx, row, row_styles = await read_queue.get()
            tasks.append(
                asyncio.create_task(
                    self.transform_row(
                        method,
                        row_idx,
                        row,
                        row_styles,
                    ),
                ),
            )

            rows_processed += 1
            if rows_processed % TRANSFORMATION_TASK_MAX_PARALLEL_LINES == 0:
                group = asyncio.gather(*tasks)
                try:
                    results = await group
                    for result in results:
                        await result_store.put(result)
                except Exception as e:
                    logger.exception('Error during applying transformations.')
                    group.cancel()
                    raise e
                tasks = []
            if rows_processed % delta == 0 or rows_processed == total_rows:
                logger.info(
                    f'Starting transformation tasks for {tfn_request["id"]}:'
                    f' {rows_processed}/{total_rows} started',
                )

        group = asyncio.gather(*tasks)
        try:
            results = await group
            for result in results:
                await result_store.put(result)
        except Exception as e:
            logger.exception('Error during applying transformations.')
            group.cancel()
            raise e

    async def transform_row(self, method, row_idx, row, row_styles):
        try:
            if ROW_DELETED_MARKER in list(row.values()):
                return RowTransformationResponse.delete()
            kwargs = {}
            if 'row_styles' in inspect.signature(method).parameters:
                kwargs['row_styles'] = row_styles
            if inspect.iscoroutinefunction(method):
                awaitable = method(row, **kwargs)
            else:
                loop = asyncio.get_running_loop()
                awaitable = loop.run_in_executor(
                    self.executor,
                    functools.partial(method, row, **kwargs),
                )
            timeout = self.config.get_timeout('row_transformation')
            response = await asyncio.wait_for(
                awaitable,
                timeout=timeout,
            )
            if not isinstance(response, RowTransformationResponse):
                raise RowTransformationError(f'invalid row tranformation response: {response}')
            if response.status == ResultType.FAIL:
                raise RowTransformationError(f'row transformation failed: {response.output}')
            return response
        except Exception as e:
            cause = (
                str(e) if not isinstance(e, asyncio.TimeoutError)
                else f'timed out after {timeout} s'
            )
            raise RowTransformationError(
                f'Error applying transformation function {method.__name__} '
                f'to row #{row_idx}: {cause}.',
            ) from e

    def write_excel(
        self, filename, result_store, total_rows, output_columns, task_data, logger, loop,
    ):
        wb = Workbook(write_only=True)

        ws = wb.create_sheet('Data')
        ws_columns = wb.create_sheet('Columns')
        ws_columns.append(['Name', 'Type', 'Nullable', 'Description', 'Precision'])
        column_keys = ['name', 'type', 'nullable', 'description', 'precision']

        column_names = []

        for column in output_columns:
            row = [column.get(key) for key in column_keys]
            ws_columns.append(row)
            column_names.append(column.get('name'))

        ws.append(column_names)

        rows_processed = 0
        delta = min(
            1 if total_rows <= 10 else round(total_rows / 10),
            REPORT_EVERY_ROW_MAX,
        )

        for _ in range(2, total_rows + 2):
            future = asyncio.run_coroutine_threadsafe(
                result_store.get(),
                loop,
            )
            response = future.result(
                timeout=self.config.env['transformation_write_queue_timeout'],
            )

            ws.append(self.generate_output_row(ws, column_names, response))
            rows_processed += 1
            if rows_processed % delta == 0 or rows_processed == total_rows:
                self.send_stat_update(task_data, rows_processed, total_rows)
                logger.info(
                    f'Writing to output file for {task_data.input.object_id}: {rows_processed}/'
                    f'{total_rows} written',
                )

        wb.save(filename)

    def generate_output_row(self, ws, column_names, response):
        row = []
        for col_name in column_names:
            if response.status == ResultType.SUCCESS:
                value = response.transformed_row.get(col_name)
                cell = WriteOnlyCell(ws=ws, value=value if value is not None else EXCEL_NULL_MARKER)
                style = response.transformed_row_styles.get(col_name, NamedStyle())
                cell.number_format = style.number_format
                cell.font = style.font
                cell.fill = style.fill
                cell.border = style.border
                row.append(cell)
            elif response.status == ResultType.SKIP:
                row.append(EXCEL_NULL_MARKER)
            elif response.status == ResultType.DELETE:
                row.append(ROW_DELETED_MARKER)
            else:
                raise RowTransformationError(
                    f'Invalid row transformation response status: {response.status}.',
                )
        return row

    def send_stat_update(self, task_data, rows_processed, total_rows):
        client = self.get_sync_client(task_data)
        ns, _ = task_data.input.event_type.split('_', 1)
        client(ns).requests[task_data.input.object_id].update(
            payload={'stats': {'rows': {'total': total_rows, 'processed': rows_processed}}},
        )

    async def send_output_file(self, task_data, batch_id, output_file, logger):
        client = self.get_client(task_data)

        fileobj = open(output_file.name, 'rb')
        fileobj.seek(0, os.SEEK_END)
        file_size = fileobj.tell()
        fileobj.seek(0)

        filename = f'{task_data.input.object_id}-out.{output_file.name.split(".")[-1]}'
        headers = {
            'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'Content-Disposition': f'attachment; filename="{filename}"',
            'Content-Length': str(file_size),
        }

        async def chunks_iterator():    # pragma: no cover
            progress = 0
            while data := fileobj.read(UPLOAD_CHUNK_SIZE):
                yield data
                progress += len(data)
                logger.debug(
                    f'Output file upload progress for {task_data.input.object_id}:'
                    f' {progress}/{file_size} bytes',
                )

        media_file = await client.ns('media').ns('folders').collection(
            'streams_batches',
        )[batch_id].action(
            'files',
        ).post(
            content=chunks_iterator(),
            headers=headers,
        )
        logger.info(
            f'Output file upload completed for {task_data.input.object_id}',
        )
        media_file_id = json.loads(media_file)['id']
        ns, _ = task_data.input.event_type.split('_', 1)
        await client(ns).requests[task_data.input.object_id].update(
            payload={'files': {'output': {'id': media_file_id}}},
        )
        await client(ns).requests[task_data.input.object_id]('process').post()

    def format_exception_message(self, e):
        if isinstance(e, asyncio.CancelledError):
            return 'cancelled'
        elif isinstance(e, asyncio.TimeoutError):
            return 'timed out'
        else:
            return str(e) or repr(e)
